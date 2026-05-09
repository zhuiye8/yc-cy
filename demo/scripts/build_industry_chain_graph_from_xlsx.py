"""Generate runtime industry-chain graph data from the 50 IPC xlsx sheets.

The browser does not read xlsx files. This script converts the local xlsx
source directory into src/platform/data/industry-chain-graph.js.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT.parent / "50条输出_IPC归集版"
DEFAULT_OUT = ROOT / "src" / "platform" / "data" / "industry-chain-graph.js"

GICS_NAMES = {
    "15101050": "专用化学品",
    "20101010": "航空航天与国防",
    "20104010": "电气部件与设备",
    "20106010": "建筑机械与重型运输设备",
    "20106020": "工业机械",
    "25101010": "汽车零部件",
    "25102010": "汽车制造",
    "35101010": "医疗保健设备",
    "35101020": "医疗保健用品",
    "35201010": "生物科技",
    "35202010": "制药",
    "45201020": "通信设备",
    "45202030": "电子设备和仪器",
    "45203015": "电子元件",
    "45301010": "半导体设备",
    "45301020": "半导体",
}


def classify_gics_group(gics_code: str) -> dict[str, str]:
    if gics_code.startswith("201010"):
        return {"key": "aerospace", "name": "航空航天与国防", "shortName": "航天", "color": "#ffb066"}
    if gics_code.startswith("201040"):
        return {"key": "electrical", "name": "电气设备", "shortName": "电气", "color": "#ffd980"}
    if gics_code.startswith("201060"):
        return {"key": "machinery", "name": "机械装备与工业设备", "shortName": "机械", "color": "#e89060"}

    prefix = gics_code[:2]
    if prefix == "15":
        return {"key": "materials", "name": "材料", "shortName": "材料", "color": "#b8aaff"}
    if prefix == "25":
        return {"key": "auto", "name": "汽车与消费品", "shortName": "汽车", "color": "#ff9bd1"}
    if prefix == "35":
        return {"key": "health", "name": "医疗健康", "shortName": "医疗", "color": "#86e4ff"}
    if prefix == "45":
        return {"key": "tech", "name": "信息技术 / 半导体", "shortName": "信息", "color": "#6bffcf"}
    return {"key": "other", "name": "其他", "shortName": "其他", "color": "#86e4ff"}


def extract_sector_list(path: Path) -> dict[tuple[str, str], dict]:
    if not path.exists():
        return {}

    text = path.read_text(encoding="utf-8")
    marker = "export const sectorList = "
    start = text.find(marker)
    if start < 0:
        return {}

    i = start + len(marker)
    while i < len(text) and text[i].isspace():
        i += 1
    if i >= len(text) or text[i] != "[":
        return {}

    depth = 0
    in_string = False
    escape = False
    end = None
    for j in range(i, len(text)):
        ch = text[j]
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                end = j + 1
                break

    if end is None:
        return {}

    try:
      rows = json.loads(text[i:end])
    except json.JSONDecodeError:
      return {}

    return {(str(row.get("gicsCode", "")), str(row.get("name", ""))): row for row in rows}


def split_cell(value) -> list[str]:
    if value is None:
        return []
    parts = re.split(r"[；;、,\n\r]+", str(value))
    result = []
    seen = set()
    for part in parts:
        item = part.strip()
        if item and item not in seen:
            result.append(item)
            seen.add(item)
    return result


def add_path(root: dict, id_prefix: str, counter: list[int], chain_name: str, path_text, leaf_name, row_meta: dict):
    if not leaf_name:
        return

    parts = [part.strip() for part in str(path_text or "").split(">") if part and part.strip()]
    if parts and (parts[0] == chain_name or parts[0] == f"{chain_name}产业链" or parts[0].endswith("产业链")):
        parts = parts[1:]
    if not parts:
        parts = ["未分组"]

    current = root
    for part in parts:
        current.setdefault("children", [])
        child = next((node for node in current["children"] if node.get("name") == part), None)
        if child is None:
            counter[0] += 1
            child = {"id": f"{id_prefix}-{counter[0]}", "name": part, "children": []}
            current["children"].append(child)
        current = child

    current.setdefault("children", [])
    counter[0] += 1
    leaf = {
        "id": f"{id_prefix}-{counter[0]}",
        "name": str(leaf_name).strip(),
    }
    if row_meta:
        leaf["meta"] = row_meta
    current["children"].append(leaf)


def split_in_thirds(items: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    if not items:
        return [], [], []
    a = (len(items) + 2) // 3
    remain = len(items) - a
    b = (remain + 1) // 2
    return items[:a], items[a:a + b], items[a + b:]


def parse_workbook(path: Path, data_key: str, chain_name: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path.name}: empty workbook")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    index = {name: i for i, name in enumerate(headers)}
    required = ["上级路径", "最末端环节"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"{path.name}: missing columns {', '.join(missing)}")

    root = {"id": f"{data_key}-r", "name": chain_name, "children": []}
    counter = [0]

    for row in rows[1:]:
        path_text = row[index["上级路径"]] if index["上级路径"] < len(row) else None
        leaf_name = row[index["最末端环节"]] if index["最末端环节"] < len(row) else None
        if not path_text or not leaf_name:
            continue

        row_meta = OrderedDict()
        for source, target in [
            ("企业关键词_主方向", "enterprisePrimary"),
            ("企业关键词_辅助", "enterpriseAuxiliary"),
            ("专利关键词_主方向", "patentPrimary"),
            ("专利关键词_辅助", "patentAuxiliary"),
            ("IPC编号_主方向", "ipcPrimary"),
            ("IPC编号_辅助", "ipcAuxiliary"),
        ]:
            values = split_cell(row[index[source]]) if source in index and index[source] < len(row) else []
            if values:
                row_meta[target] = values
        if "覆盖强度" in index and index["覆盖强度"] < len(row) and row[index["覆盖强度"]]:
            row_meta["coverageStrength"] = str(row[index["覆盖强度"]]).strip()

        add_path(root, data_key, counter, chain_name, path_text, leaf_name, dict(row_meta))

    return root


def emit_js(graph: OrderedDict, sector_list: list[dict], source_dir: Path) -> str:
    return (
        "/**\n"
        " * 产业链图谱数据 — 由 scripts/build_industry_chain_graph_from_xlsx.py 从 50 个 IPC 归集 xlsx 表单生成。\n"
        " *\n"
        f" * 转换源：{source_dir.as_posix()}\n"
        " * 运行时只读取本文件，不直接依赖 xlsx；如表单更新，重新运行脚本刷新。\n"
        " */\n\n"
        f"export const industryChainGraphData = {json.dumps(graph, ensure_ascii=False, indent=2)}\n\n"
        f"export const sectorList = {json.dumps(sector_list, ensure_ascii=False, indent=2)}\n"
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    source = args.source.resolve()
    out = args.out.resolve()
    if not source.exists():
        print(f"source directory not found: {source}", file=sys.stderr)
        return 1

    files = [p for p in sorted(source.glob("*.xlsx")) if not p.name.startswith("~$")]
    if not files:
        print(f"no xlsx files found: {source}", file=sys.stderr)
        return 1

    existing_meta = extract_sector_list(out)
    graph = OrderedDict()
    sector_list = []
    per_gics_count: dict[str, int] = {}

    for file_path in files:
        match = re.match(r"^(\d{8})_(.+?)_.*$", file_path.stem)
        if not match:
            raise ValueError(f"unexpected xlsx filename: {file_path.name}")

        gics_code, chain_name = match.groups()
        index = per_gics_count.get(gics_code, 0)
        per_gics_count[gics_code] = index + 1
        data_key = f"c_{gics_code}_{index}"

        root = parse_workbook(file_path, data_key, chain_name)
        upstream, midstream, downstream = split_in_thirds(root.get("children", []))
        graph[data_key] = OrderedDict()
        for key, label, children in [
            ("upstream", "上游", upstream),
            ("midstream", "中游", midstream),
            ("downstream", "下游", downstream),
        ]:
            if children:
                graph[data_key][key] = {
                    "label": label,
                    "root": {"id": f"{data_key}-{key[:2]}", "name": chain_name, "children": children},
                }

        meta = existing_meta.get((gics_code, chain_name), {})
        group = classify_gics_group(gics_code)
        sector_list.append({
            "dataKey": data_key,
            "name": chain_name,
            "gicsCode": gics_code,
            "gicsName": meta.get("gicsName") or GICS_NAMES.get(gics_code, gics_code),
            "groupKey": meta.get("groupKey") or group["key"],
            "groupName": meta.get("groupName") or group["name"],
            "groupShort": meta.get("groupShort") or group["shortName"],
            "colorHex": meta.get("colorHex") or group["color"],
            "position": meta.get("position") or {"x": 0, "y": 0, "z": 0},
        })

    out.write_text(emit_js(graph, sector_list, source), encoding="utf-8")
    leaf_count = sum(1 for _ in iter_leaves(graph))
    print(f"wrote {out}")
    print(f"source xlsx files: {len(files)}")
    print(f"sector count: {len(sector_list)}")
    print(f"leaf count: {leaf_count}")
    return 0


def iter_leaves(graph: OrderedDict):
    def walk(node):
        children = node.get("children") or []
        if not children:
            yield node
        for child in children:
            yield from walk(child)

    for segments in graph.values():
        for segment in segments.values():
            yield from walk(segment["root"])


if __name__ == "__main__":
    raise SystemExit(main())
